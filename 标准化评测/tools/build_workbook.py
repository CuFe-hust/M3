#!/usr/bin/env python3
"""Build 评测表.xlsx from history CSVs + registry/metrics.yaml (Task 8).

Usage:
    python tools/build_workbook.py <history_root> <registry_path> <output_xlsx>
"""

from __future__ import annotations

import csv
import json
import re
import sys
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.cell import range_boundaries

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="Microsoft YaHei", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="1F4E79")
WARNING_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GOOD_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

# Formula error patterns (case-insensitive)
_FORMULA_ERROR_RE = re.compile(
    r"#REF!|#DIV/0!|#VALUE!|#NAME\?|#N/A|#NULL!|#NUM!", re.IGNORECASE
)


def _safe_float(value: Any) -> float | None:
    """Convert a CSV cell value to float, or None if empty/invalid."""
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _display_value(canonical: float | None, multiplier: float) -> str | None:
    """Format a canonical value for display (e.g. ratio * 100)."""
    if canonical is None:
        return None
    return str(round(canonical * multiplier, 4))


# ---------------------------------------------------------------------------
# Sheet 1: 使用说明
# ---------------------------------------------------------------------------
def _sheet_instructions(ws) -> None:
    ws.title = "使用说明"
    ws.sheet_properties.tabColor = "1F4E79"

    lines = [
        ["M3-RS 标准化评测系统 —— 核心《评测表》"],
        [""],
        ["填写规则"],
        ["1. 本工作簿由 build_workbook.py 从 history/ CSV 文件和 registry/metrics.yaml 自动生成。"],
        ["2. 不得手工修改数据单元格；人工备注请写在「协议与质检」工作表的备注栏。"],
        ["3. 比率类指标 (canonical_unit=ratio) 以 0-1 入库，显示时乘 100（百分比）。"],
        ["4. CIDEr/CIDEr-D 以 COCO 原始分数入库，显示时乘 100。"],
        ["5. 缺失值留空或记 N/A，禁止以 0 代替。"],
        ["6. 仅 mode=full 且 status=complete 的正式运行进入「核心评测表」和「对比看板」。"],
        ["7. 冒烟测试 (mode=smoke) 和未完成运行仅记录在「指标长表」和「数据覆盖」中。"],
        [""],
        ["指标单位说明"],
        ["- ratio: 已乘 100 显示为 %。"],
        ["- points: 竞赛评分，原始单位。"],
        ["- milliseconds (ms): 延迟。"],
        ["- count: 计数单位。"],
        ["- words: 词数。"],
        ["- score / paper points: CIDEr 类分数。"],
        ["- degrees: 角度误差。"],
        [""],
        ["运行流程"],
        ["1. 预检: python -m m3rs_eval doctor --config configs/server.yaml"],
        ["2. 冒烟: python -m m3rs_eval run --config configs/server.yaml --mode smoke --limit 2"],
        ["3. 全量: python -m m3rs_eval run --config configs/server.yaml --mode full"],
        ["4. 重建评测表: python -m m3rs_eval rebuild-table"],
        ["5. 生成报告: python -m m3rs_eval prepare-report --run-id <RUN_ID>"],
        [""],
        ["风险提示"],
        ["- 本系统不修改数据集标准答案或跳过失败样本。"],
        ["- Full 与 Lite、英文与中文数据禁止混列为同一平均分。"],
        ["- LLM Judge 结果仅作辅助，不计入竞赛主评分。"],
        ["- 未经协议检查的运行不得与其他正式运行直接比较。"],
        [""],
        [f"生成时间: 由 build_workbook.py 运行时记录"],
    ]

    for row_idx, line in enumerate(lines, 1):
        cell = ws.cell(row=row_idx, column=1, value=line[0])
        if row_idx == 1:
            cell.font = Font(name="Microsoft YaHei", bold=True, size=14, color="1F4E79")
        elif line[0].startswith("填写规则") or line[0].startswith("指标单位说明") or \
             line[0].startswith("运行流程") or line[0].startswith("风险提示"):
            cell.font = SECTION_FONT
        else:
            cell.font = DATA_FONT

    ws.column_dimensions["A"].width = 80


# ---------------------------------------------------------------------------
# Sheet 2: 核心评测表
# ---------------------------------------------------------------------------
def _sheet_core_scorecard(ws, runs: list[dict], metrics_by_run: dict, registry: dict, reg_ids: set) -> None:
    ws.title = "核心评测表"
    ws.sheet_properties.tabColor = "C00000"

    # Build header
    headers = [
        # 运行身份
        "run_id", "system_version", "model_name", "training_data_version",
        "code_commit", "protocol_id", "recorded_at",
        # 完整性
        "status", "total_failures", "failure_rate_display", "quality_comparable",
        # VRSBench Caption (display multiplier *100 for ratios)
        "VRS_Caption_BLEU4", "VRS_Caption_METEOR", "VRS_Caption_ROUGEL",
        # VRSBench Grounding
        "VRS_Ground_UniqAcc05", "VRS_Ground_NonUniqAcc05", "VRS_Ground_AllAcc05",
        # VRSBench VQA
        "VRS_VQA_AllAcc",
        # MME-RS
        "MME_Color", "MME_Count", "MME_Position", "MME_Avg", "MME_Avg_C",
        # XLRS-Bench Full
        "XLRS_EN_VQA_AvgL2", "XLRS_EN_Caption_BLEU4", "XLRS_EN_Ground_Acc05",
        # LEVIR-CC
        "LEVIR_All_BLEU4", "LEVIR_All_METEOR", "LEVIR_All_ROUGEL",
        # 资源
        "total_run_seconds", "e2e_p50_ms", "e2e_p95_ms",
        "throughput_samples_per_s", "total_params_b", "model_storage_gib",
        # 产物
        "artifact_path",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Only eligible runs
    eligible = [r for r in runs if r.get("eligible_for_history", "").strip().lower() == "true"]
    row_idx = 2
    for run in eligible:
        rid = run["run_id"]
        mets = metrics_by_run.get(rid, {})

        def _disp(metric_id: str) -> str | None:
            val = mets.get(metric_id)
            if val is None:
                return None
            mult = registry.get(metric_id, {}).get("display_multiplier", 1)
            return _display_value(val, float(mult))

        total_fail = sum(
            1 for m_id, v in mets.items()
            if m_id in reg_ids and _safe_float(registry.get(m_id, {}).get("n_failures", 0)) is not None
        )
        # approximate failure_rate from metrics
        total_samples = sum(
            int(m.get("n_samples", 0) or 0)
            for rmid, m in metrics_by_run.items() if rmid == rid
            for m_id, m in [("_", m)] if m_id in reg_ids
        )
        failure_rate = total_fail / total_samples if total_samples > 0 else None
        failure_rate_disp = _display_value(failure_rate, 100) if failure_rate is not None else None

        values = [
            rid,
            run.get("system_version", ""),
            run.get("model_name", ""),
            run.get("training_data_version", ""),
            run.get("code_commit", ""),
            run.get("protocol_id", ""),
            run.get("recorded_at", ""),
            run.get("status", ""),
            total_fail,
            failure_rate_disp,
            "TBD",
            _disp("vrs.caption.bleu_4"),
            _disp("vrs.caption.meteor"),
            _disp("vrs.caption.rouge_l"),
            _disp("vrs.grounding.hbb.unique.acc_0_5"),
            _disp("vrs.grounding.hbb.non_unique.acc_0_5"),
            _disp("vrs.grounding.hbb.all.acc_0_5"),
            _disp("vrs.vqa.acc.all"),
            _disp("mme_rs.acc.color"),
            _disp("mme_rs.acc.count"),
            _disp("mme_rs.acc.position"),
            _disp("mme_rs.avg"),
            _disp("mme_rs.avg_c"),
            _disp("xlrs.vqa.en.paper_avg_l2"),
            _disp("xlrs.caption.en.bleu_4"),
            _disp("xlrs.grounding.en.all.acc_0_5"),
            _disp("levir.caption.all.bleu_4"),
            _disp("levir.caption.all.meteor"),
            _disp("levir.caption.all.rouge_l"),
            _safe_float(run.get("total_run_seconds")),
            _safe_float(run.get("e2e_p50_ms")),
            _safe_float(run.get("e2e_p95_ms")),
            _safe_float(run.get("throughput_samples_per_s")),
            _safe_float(run.get("total_params_b")),
            _safe_float(run.get("model_storage_gib")),
            run.get("artifact_path", ""),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val if val is not None else "")
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col == 1:
                cell.alignment = Alignment(horizontal="left")
        row_idx += 1

    # Auto-filter and freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"
    ws.freeze_panes = "A2"

    # Column widths
    col_widths = [45, 16, 20, 14, 14, 22, 20, 10, 10, 12, 14] + [16] * 20 + [12] * 6 + [30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Sheet 3: 指标长表
# ---------------------------------------------------------------------------
def _sheet_metrics_long(ws, metrics_rows: list[dict]) -> None:
    ws.title = "指标长表"
    ws.sheet_properties.tabColor = "ED7D31"

    headers = [
        "run_id", "metric_id", "value_canonical", "value_display",
        "n_samples", "n_failures", "ci95_low", "ci95_high",
        "dataset", "task", "slice", "language",
        "availability", "provenance", "recorded_at",
        "protocol_id", "benchmark_version", "notes",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, mrow in enumerate(metrics_rows, 2):
        canonical = _safe_float(mrow.get("value_canonical"))
        multiplier = 100 if mrow.get("canonical_unit", "ratio") == "ratio" else 1
        disp = _display_value(canonical, multiplier) if canonical is not None else ""

        vals = [
            mrow.get("run_id", ""),
            mrow.get("metric_id", ""),
            canonical,
            disp,
            _safe_float(mrow.get("n_samples")),
            _safe_float(mrow.get("n_failures")),
            _safe_float(mrow.get("ci95_low")),
            _safe_float(mrow.get("ci95_high")),
            mrow.get("dataset", ""),
            mrow.get("task", ""),
            mrow.get("slice", ""),
            mrow.get("language", ""),
            mrow.get("availability", ""),
            mrow.get("provenance", ""),
            mrow.get("recorded_at", ""),
            mrow.get("protocol_id", ""),
            mrow.get("benchmark_version", ""),
            mrow.get("notes", ""),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val if val is not None else "")
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

    last_row = len(metrics_rows) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"
    ws.freeze_panes = "A2"

    col_widths = [45, 40, 14, 14, 10, 10, 10, 10, 16, 22, 18, 8, 12, 16, 20, 18, 18, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Sheet 4: 运行元数据
# ---------------------------------------------------------------------------
def _sheet_run_metadata(ws, runs: list[dict]) -> None:
    ws.title = "运行元数据"
    ws.sheet_properties.tabColor = "4472C4"

    metadata_fields = [
        "run_id", "status", "mode", "eligible_for_history", "recorded_at",
        "protocol_id", "config_hash", "system_version", "model_name",
        "total_run_seconds", "e2e_p50_ms", "e2e_p95_ms",
        "throughput_samples_per_s", "total_params_b", "model_storage_gib",
        "hardware_summary", "code_commit", "seed", "artifact_path", "notes",
    ]
    for col, h in enumerate(metadata_fields, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, run in enumerate(runs, 2):
        for col, field in enumerate(metadata_fields, 1):
            val = run.get(field, "")
            cell = ws.cell(row=row_idx, column=col, value=val if val is not None else "")
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

    last_row = len(runs) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(metadata_fields))}{last_row}"
    ws.freeze_panes = "A2"

    col_widths = [45, 10, 8, 14, 20, 18, 14, 14, 20, 14, 10, 10, 14, 12, 14, 18, 16, 6, 42, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Sheet 5: 数据覆盖
# ---------------------------------------------------------------------------
def _sheet_coverage(ws, coverage_rows: list[dict]) -> None:
    ws.title = "数据覆盖"
    ws.sheet_properties.tabColor = "70AD47"

    headers = ["run_id", "dataset", "task", "expected", "requested", "predicted", "failed", "status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, crow in enumerate(coverage_rows, 2):
        for col, field in enumerate(headers, 1):
            val = crow.get(field, "")
            cell = ws.cell(row=row_idx, column=col, value=val if val is not None else "")
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            # Color-code status
            if field == "status":
                if val == "complete":
                    cell.fill = GOOD_FILL
                elif val == "incomplete":
                    cell.fill = WARNING_FILL

    last_row = len(coverage_rows) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"
    ws.freeze_panes = "A2"

    col_widths = [45, 18, 24, 10, 10, 10, 8, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Sheet 6: 对比看板
# ---------------------------------------------------------------------------
def _sheet_comparison_dashboard(
    ws, runs: list[dict], metrics_by_run: dict, registry_data: dict
) -> None:
    ws.title = "对比看板"
    ws.sheet_properties.tabColor = "FFC000"

    eligible = [r for r in runs if r.get("eligible_for_history", "").strip().lower() == "true"]
    if len(eligible) < 2:
        ws.cell(row=1, column=1, value="需要至少两次 eligible 正式运行才能生成对比看板。").font = DATA_FONT
        return

    # Use the two eligible runs as candidate (most recent) and baseline (older)
    runs_sorted = sorted(eligible, key=lambda r: r.get("recorded_at", ""))
    baseline = runs_sorted[0]
    candidate = runs_sorted[-1]  # most recent

    b_id = baseline["run_id"]
    c_id = candidate["run_id"]
    b_mets = metrics_by_run.get(b_id, {})
    c_mets = metrics_by_run.get(c_id, {})

    # Key metrics for comparison
    key_metrics = [
        ("VRSBench Caption BLEU-4", "vrs.caption.bleu_4", "higher", 100),
        ("VRSBench Caption METEOR", "vrs.caption.meteor", "higher", 100),
        ("VRSBench Caption ROUGE-L", "vrs.caption.rouge_l", "higher", 100),
        ("VRSBench Ground Unique Acc@0.5", "vrs.grounding.hbb.unique.acc_0_5", "higher", 100),
        ("VRSBench Ground Non-Unique Acc@0.5", "vrs.grounding.hbb.non_unique.acc_0_5", "higher", 100),
        ("VRSBench Ground All Acc@0.5", "vrs.grounding.hbb.all.acc_0_5", "higher", 100),
        ("VRSBench VQA All Accuracy", "vrs.vqa.acc.all", "higher", 100),
        ("MME-RS Color", "mme_rs.acc.color", "higher", 100),
        ("MME-RS Count", "mme_rs.acc.count", "higher", 100),
        ("MME-RS Position", "mme_rs.acc.position", "higher", 100),
        ("MME-RS Avg", "mme_rs.avg", "higher", 100),
        ("MME-RS Avg-C", "mme_rs.avg_c", "higher", 100),
        ("XLRS EN VQA Avg L2", "xlrs.vqa.en.paper_avg_l2", "higher", 100),
        ("XLRS EN Caption BLEU-4", "xlrs.caption.en.bleu_4", "higher", 100),
        ("XLRS EN Ground Acc@0.5", "xlrs.grounding.en.all.acc_0_5", "higher", 100),
        ("LEVIR-CC All BLEU-4", "levir.caption.all.bleu_4", "higher", 100),
    ]

    headers = [
        "指标", "方向", "基线值", "候选值", "原始变化量",
        "方向调整改善量", "可比性", "基线run_id", "候选run_id",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, (label, m_id, direction, disp_mult) in enumerate(key_metrics, 2):
        b_val = b_mets.get(m_id)
        c_val = c_mets.get(m_id)

        b_disp = _display_value(b_val, disp_mult) if b_val is not None else ""
        c_disp = _display_value(c_val, disp_mult) if c_val is not None else ""

        raw_delta = None
        adj_improvement = ""
        if b_val is not None and c_val is not None and b_val != 0:
            raw_delta = (c_val - b_val) / abs(b_val) * 100
            if direction == "higher":
                adj_improvement = f"+{raw_delta:.2f}%" if raw_delta >= 0 else f"{raw_delta:.2f}%"
            elif direction == "lower":
                adj_improvement = f"+{abs(raw_delta):.2f}%" if raw_delta <= 0 else f"{-raw_delta:.2f}%"
        else:
            raw_delta = None

        values = [
            label, direction,
            b_disp if b_disp else "N/A",
            c_disp if c_disp else "N/A",
            f"{raw_delta:.2f}%" if raw_delta is not None else "N/A",
            adj_improvement or "N/A",
            "TBD",
            b_id, c_id,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

    last_row = len(key_metrics) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"
    ws.freeze_panes = "A2"

    col_widths = [32, 8, 12, 12, 14, 16, 10, 45, 45]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Sheet 7: 指标字典
# ---------------------------------------------------------------------------
def _sheet_metric_dictionary(ws, registry_metrics: list[dict]) -> None:
    ws.title = "指标字典"
    ws.sheet_properties.tabColor = "7030A0"

    headers = [
        "metric_id", "name_zh", "authority", "direction",
        "canonical_unit", "display_multiplier", "display_unit",
        "dataset", "task", "slice", "language",
        "formula_short", "availability", "caveats",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, metric in enumerate(registry_metrics, 2):
        formula = metric.get("formula", "")
        if len(formula) > 120:
            formula = formula[:117] + "..."

        values = [
            metric.get("metric_id", ""),
            metric.get("name_zh", ""),
            metric.get("authority", ""),
            metric.get("direction", ""),
            metric.get("canonical_unit", ""),
            metric.get("display_multiplier", 1),
            metric.get("display_unit", ""),
            metric.get("dataset", ""),
            metric.get("task", ""),
            metric.get("slice", ""),
            metric.get("language", ""),
            formula,
            metric.get("availability", ""),
            metric.get("caveats", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val if val is not None else "")
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

    last_row = len(registry_metrics) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"
    ws.freeze_panes = "A2"

    col_widths = [42, 30, 8, 10, 14, 14, 12, 18, 26, 22, 8, 60, 12, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Sheet 8: 协议与质检
# ---------------------------------------------------------------------------
def _sheet_protocol_qc(ws, registry_data: dict, runs: list[dict]) -> None:
    ws.title = "协议与质检"
    ws.sheet_properties.tabColor = "BF8F00"

    # Sources section
    row = 1
    ws.cell(row=row, column=1, value="权威来源").font = SECTION_FONT
    row += 1
    source_headers = ["id", "title", "type", "location", "key_sections", "note"]
    for col, h in enumerate(source_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    sources = registry_data.get("sources", [])
    for src in sources:
        for col, key in enumerate(source_headers, 1):
            cell = ws.cell(row=row, column=col, value=src.get(key, ""))
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="协议版本信息").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"registry_version: {registry_data.get('registry_version', 'N/A')}").font = DATA_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"generated_on: {registry_data.get('generated_on', 'N/A')}").font = DATA_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"project: {registry_data.get('project', 'N/A')}").font = DATA_FONT

    row += 2
    ws.cell(row=row, column=1, value="运行协议检查").font = SECTION_FONT
    row += 1
    protocol_headers = ["run_id", "protocol_id", "status", "mode", "eligible_for_history"]
    for col, h in enumerate(protocol_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    for run in runs:
        for col, key in enumerate(protocol_headers, 1):
            cell = ws.cell(row=row, column=col, value=run.get(key, ""))
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="质检告警与备注").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value="（人工修改说明、已知冲突、缺失项请填写在此区域。）").font = DATA_FONT
    row += 1
    ws.cell(row=row, column=1, value="").font = DATA_FONT
    row += 1
    ws.cell(row=row, column=1, value="自动质检: 已检查所有指标 metric_id 均在注册表中。").font = DATA_FONT

    # Column widths
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 60


# ---------------------------------------------------------------------------
# Workbook verification
# ---------------------------------------------------------------------------
def _verify_workbook(wb: Workbook, expected_sheets: int = 8, metrics_rows: int = 0,
                     coverage_rows: int = 0, eligible_runs: int = 0,
                     registry_metrics: int = 194) -> list[str]:
    """Verify workbook integrity. Returns list of error messages."""
    errors: list[str] = []

    if len(wb.sheetnames) != expected_sheets:
        errors.append(
            f"Expected {expected_sheets} sheets, got {len(wb.sheetnames)}: {wb.sheetnames}"
        )

    sheet_names_expected = [
        "使用说明", "核心评测表", "指标长表", "运行元数据",
        "数据覆盖", "对比看板", "指标字典", "协议与质检",
    ]
    missing = set(sheet_names_expected) - set(wb.sheetnames)
    if missing:
        errors.append(f"Missing sheets: {missing}")

    # Check metrics_long row count
    if "指标长表" in wb.sheetnames:
        ws = wb["指标长表"]
        data_rows = ws.max_row - 1  # minus header
        if metrics_rows > 0 and data_rows != metrics_rows:
            errors.append(
                f"指标长表: expected {metrics_rows} data rows, got {data_rows}"
            )

    # Check coverage row count
    if "数据覆盖" in wb.sheetnames:
        ws_cov = wb["数据覆盖"]
        cov_rows = ws_cov.max_row - 1
        if coverage_rows > 0 and cov_rows != coverage_rows:
            errors.append(
                f"数据覆盖: expected {coverage_rows} data rows, got {cov_rows}"
            )

    # Check metric dictionary row count
    if "指标字典" in wb.sheetnames:
        ws_dict = wb["指标字典"]
        dict_rows = ws_dict.max_row - 1
        if dict_rows != registry_metrics:
            errors.append(
                f"指标字典: expected {registry_metrics} metrics, got {dict_rows}"
            )

    # Scan all cells for formula errors
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and _FORMULA_ERROR_RE.search(cell.value):
                    errors.append(
                        f"[{ws_name}] Cell {cell.coordinate}: formula error '{cell.value}'"
                    )

    return errors


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------
def main(history_root: str, registry_path: str, output_path: str) -> int:
    history_dir = Path(history_root)
    registry_file = Path(registry_path)
    output_xlsx = Path(output_path)

    # Validate inputs
    if not history_dir.is_dir():
        print(f"ERROR: history_root is not a directory: {history_dir}", file=sys.stderr)
        return 1
    if not registry_file.is_file():
        print(f"ERROR: registry_path is not a file: {registry_file}", file=sys.stderr)
        return 1

    # Load registry
    print(f"Loading registry from {registry_file} ...")
    with registry_file.open("r", encoding="utf-8") as fh:
        registry_data = yaml.safe_load(fh)
    registry_metrics = registry_data.get("metrics", [])
    registry_by_id: dict[str, dict] = {m["metric_id"]: m for m in registry_metrics}
    reg_ids = set(registry_by_id.keys())
    print(f"  Loaded {len(registry_metrics)} metrics from registry.")

    # Load history CSVs
    print(f"Loading history CSVs from {history_dir} ...")
    runs_csv = history_dir / "runs.csv"
    metrics_csv = history_dir / "metrics_long.csv"
    coverage_csv = history_dir / "coverage.csv"

    if not runs_csv.is_file():
        print(f"ERROR: runs.csv not found at {runs_csv}", file=sys.stderr)
        return 2
    if not metrics_csv.is_file():
        print(f"ERROR: metrics_long.csv not found at {metrics_csv}", file=sys.stderr)
        return 2
    if not coverage_csv.is_file():
        print(f"ERROR: coverage.csv not found at {coverage_csv}", file=sys.stderr)
        return 2

    # Read runs
    with runs_csv.open("r", encoding="utf-8-sig", newline="") as fh:
        runs = list(csv.DictReader(fh))
    print(f"  Loaded {len(runs)} runs.")

    # Read metrics_long
    with metrics_csv.open("r", encoding="utf-8-sig", newline="") as fh:
        metrics_rows = list(csv.DictReader(fh))
    print(f"  Loaded {len(metrics_rows)} metric records.")

    # Read coverage
    with coverage_csv.open("r", encoding="utf-8-sig", newline="") as fh:
        coverage_rows = list(csv.DictReader(fh))
    print(f"  Loaded {len(coverage_rows)} coverage records.")

    # Validate metric IDs
    missing_ids: set[str] = set()
    for mrow in metrics_rows:
        mid = mrow.get("metric_id", "")
        if mid and mid not in reg_ids:
            missing_ids.add(mid)

    if missing_ids:
        print(
            f"ERROR: {len(missing_ids)} metric_id(s) in history not found in registry: {sorted(missing_ids)}",
            file=sys.stderr,
        )
        return 3

    # Build metrics_by_run lookup
    metrics_by_run: dict[str, dict[str, float]] = {}
    for mrow in metrics_rows:
        rid = mrow["run_id"]
        if rid not in metrics_by_run:
            metrics_by_run[rid] = {}
        val = _safe_float(mrow.get("value_canonical"))
        if val is not None:
            metrics_by_run[rid][mrow["metric_id"]] = val

    # Count eligible runs
    eligible = [r for r in runs if r.get("eligible_for_history", "").strip().lower() == "true"]
    print(f"  Eligible (formal) runs: {len(eligible)} of {len(runs)} total.")

    # Build workbook
    print("\nBuilding workbook ...")
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1: 使用说明
    ws1 = wb.create_sheet("使用说明")
    _sheet_instructions(ws1)

    # Sheet 2: 核心评测表
    ws2 = wb.create_sheet("核心评测表")
    _sheet_core_scorecard(ws2, runs, metrics_by_run, registry_by_id, reg_ids)

    # Sheet 3: 指标长表
    ws3 = wb.create_sheet("指标长表")
    _sheet_metrics_long(ws3, metrics_rows)

    # Sheet 4: 运行元数据
    ws4 = wb.create_sheet("运行元数据")
    _sheet_run_metadata(ws4, runs)

    # Sheet 5: 数据覆盖
    ws5 = wb.create_sheet("数据覆盖")
    _sheet_coverage(ws5, coverage_rows)

    # Sheet 6: 对比看板
    ws6 = wb.create_sheet("对比看板")
    _sheet_comparison_dashboard(ws6, runs, metrics_by_run, registry_data)

    # Sheet 7: 指标字典
    ws7 = wb.create_sheet("指标字典")
    _sheet_metric_dictionary(ws7, registry_metrics)

    # Sheet 8: 协议与质检
    ws8 = wb.create_sheet("协议与质检")
    _sheet_protocol_qc(ws8, registry_data, runs)

    # Verify
    print("\nVerifying workbook ...")
    errors = _verify_workbook(
        wb,
        expected_sheets=8,
        metrics_rows=len(metrics_rows),
        coverage_rows=len(coverage_rows),
        eligible_runs=len(eligible),
        registry_metrics=len(registry_metrics),
    )

    if errors:
        for err in errors:
            print(f"  VERIFICATION ERROR: {err}", file=sys.stderr)
        print(f"\nVerification FAILED: {len(errors)} error(s).", file=sys.stderr)
        return 4

    print("  Verification PASSED: all checks OK.")

    # Save
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_xlsx))
    print(f"\nWorkbook saved to: {output_xlsx}")
    print(f"  Sheets: {wb.sheetnames}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print(
            "Usage: python tools/build_workbook.py <history_root> <registry_path> <output_xlsx>",
            file=sys.stderr,
        )
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2], sys.argv[3]))
